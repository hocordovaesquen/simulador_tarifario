# app.py
# ==================================================================================
# Simulador Tarifario RV — v7 (Excel‑driven, robust totals-by-components)
# ----------------------------------------------------------------------------------
# - Edita tramos en "1. Parametros" (columna R+ T:AE filas 117–160) y los escribe al XLSX.
# - Recalcula con xlcalculator (si está disponible). Si no, usa valores cacheados (data_only)
#   y como último recurso hace fallback a 0/NA sin romper.
# - Lee "A.3 Negociación" fila por fila. Si no encuentra "Ingreso Total Real/Proyectado",
#   calcula el total sumando Acceso + Transacción + Perfiles (Real/Proy) por fila.
# - Asigna País a cada cliente por el bloque (BCS → Chile, BVL → Perú, BVC → Colombia).
# - Maneja celdas mergeadas al escribir (safe_set).
#
# Requisitos:
#   pip install streamlit pandas numpy openpyxl plotly xlcalculator
# ==================================================================================

import io
import math
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from io import BytesIO
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.cell.cell import MergedCell

try:
    from xlcalculator import ModelCompiler, Model
    XL_OK = True
except Exception:
    XL_OK = False

st.set_page_config(page_title="Simulador Tarifario RV – Excel‑driven (v7)", page_icon="📊", layout="wide")

# --------------------------------------
# Helpers
# --------------------------------------
def is_num(x):
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))

def safe_float(x, default=0.0):
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return default
        if isinstance(x, str):
            x = x.replace("$","").replace(",","").replace("\xa0","").strip()
        return float(x)
    except:
        return default

def calc_bps(ingreso, monto):
    if monto <= 0:
        return 0.0
    return (ingreso / monto) * 10000.0

def norm(s: str) -> str:
    if s is None: return ""
    s = str(s).lower()
    trans = str.maketrans("áéíóúüñ", "aeiouun")
    return s.translate(trans)

# --------------------------------------
# Workbook I/O
# --------------------------------------
PARAM_SHEET = "1. Parametros"
NEG_SHEET   = "A.3 Negociación"
MAP_RIGHT = {"Colombia": ("T","U","V","W"), "Perú": ("X","Y","Z","AA"), "Chile": ("AB","AC","AD","AE")}
ROWS_PANT   = list(range(117, 135))
ROWS_TRX    = list(range(134, 145))
ROWS_DMA    = list(range(146, 160))
DESC_BCS_POS= ("C", 114)

def load_workbook(bytes_data: bytes):
    return openpyxl.load_workbook(io.BytesIO(bytes_data), data_only=False)

def safe_set(ws, coord: str, value):
    """Escribe respetando celdas merged (anchor)."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    cell.value = value

def read_block_from_wb(wb, rows, fields=("min","max","bps","fijo")):
    ws = wb[PARAM_SHEET]
    out = {k: [] for k in MAP_RIGHT}
    for pais, (cmin, cmax, cvar, cfijo) in MAP_RIGHT.items():
        for rr in rows:
            mn = ws[f"{cmin}{rr}"].value
            mx = ws[f"{cmax}{rr}"].value
            var = ws[f"{cvar}{rr}"].value
            fijo= ws[f"{cfijo}{rr}"].value
            if any(is_num(v) for v in [mn, mx, var, fijo]):
                out[pais].append({fields[0]: safe_float(mn, 0.0), fields[1]: safe_float(mx, float("inf")), fields[2]: safe_float(var, 0.0), fields[3]: safe_float(fijo, 0.0)})
    return out

def read_params_from_wb(wb):
    ws = wb[PARAM_SHEET]
    desc = safe_float(ws[f"{DESC_BCS_POS[0]}{DESC_BCS_POS[1]}"].value, 0.15)
    trans = read_block_from_wb(wb, ROWS_TRX, ("min","max","bps","fijo"))
    dma   = read_block_from_wb(wb, ROWS_DMA, ("min","max","bps","fijo"))
    pant  = read_block_from_wb(wb, ROWS_PANT, ("min","max","var","fija"))
    return {"desc_bcs":desc, "transaccion":trans, "dma":dma, "pantallas":pant}

def write_params_to_wb(wb, params: Dict):
    ws = wb[PARAM_SHEET]
    desc = params.get("desc_bcs", None)
    if desc is not None:
        c, r = DESC_BCS_POS
        safe_set(ws, f"{c}{r}", float(desc))
    def write_block(rows: List[int], data_key: str, field_names: Tuple[str,str,str,str]):
        block = params.get(data_key, {})
        for pais, cols in MAP_RIGHT.items():
            cmin, cmax, cvar, cfijo = cols
            lst = block.get(pais, [])
            for i, rr in enumerate(rows):
                payload = lst[i] if i < len(lst) else None
                vmin  = safe_float(payload.get(field_names[0], payload.get("min", 0))) if payload else 0.0
                vmax  = safe_float(payload.get(field_names[1], payload.get("max", 0))) if payload else 0.0
                vvar  = safe_float(payload.get(field_names[2], payload.get("bps", payload.get("var", 0)))) if payload else 0.0
                vfijo = safe_float(payload.get(field_names[3], payload.get("fijo", payload.get("fija", 0)))) if payload else 0.0
                safe_set(ws, f"{cmin}{rr}", vmin); safe_set(ws, f"{cmax}{rr}", vmax)
                safe_set(ws, f"{cvar}{rr}", vvar); safe_set(ws, f"{cfijo}{rr}", vfijo)
    write_block(ROWS_TRX, "transaccion", ("min","max","bps","fijo"))
    write_block(ROWS_DMA, "dma", ("min","max","bps","fijo"))
    write_block(ROWS_PANT, "pantallas", ("min","max","var","fija"))
    return wb

# --------------------------------------
# A.3 header location and columns
# --------------------------------------
def find_header_row(ws):
    max_row = min(120, ws.max_row)
    for r in range(1, max_row+1):
        row_norm = [norm(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        if any(v.startswith("corredor") or v.startswith("cliente") for v in row_norm):
            return r
    return 9

def build_super_sub(ws, h):
    sup = {c: ws.cell(h, c).value for c in range(1, ws.max_column+1)}
    sub = {c: ws.cell(h+1, c).value for c in range(1, ws.max_column+1)}
    return sup, sub

def match_tokens(text, tokens):
    t = norm(text)
    return all(tok in t for tok in tokens)

def find_col(ws, sup, sub, super_tokens, sub_tokens_candidates):
    """Busca una columna cuya cabecera superior contenga super_tokens y la inferior uno de los sub sets"""
    for subs in sub_tokens_candidates:
        for c in range(1, ws.max_column+1):
            if match_tokens(sup.get(c,""), super_tokens) and match_tokens(sub.get(c,""), subs):
                return c
    # Si no se encontró, permitir match por cualquiera de los dos niveles
    for subs in sub_tokens_candidates:
        for c in range(1, ws.max_column+1):
            if match_tokens(sup.get(c,"") + " " + sub.get(c,""), super_tokens + subs):
                return c
    return None

def locate_columns(ws):
    h = find_header_row(ws)
    sup, sub = build_super_sub(ws, h)
    # Básicos
    cliente_c = None; monto_c = None
    for c in range(1, ws.max_column+1):
        sup_t = norm(sup.get(c,"")); sub_t = norm(sub.get(c,""))
        if cliente_c is None and ("corredor" in sup_t or "cliente" in sup_t or "corredor" in sub_t or "cliente" in sub_t):
            cliente_c = c
        if monto_c is None and ("monto" in sup_t and "neg" in sup_t) or ("monto" in sub_t and "neg" in sub_t) or ("monto" in sup_t and "usd" in sup_t):
            monto_c = c

    # Totales
    real_c = find_col(ws, sup, sub, ["ingreso","total"], [["real"], ["real ","real"]]) \
             or find_col(ws, sup, sub, ["ingresos","total"], [["real"]])
    proy_c = find_col(ws, sup, sub, ["ingreso","total"], [["proy"], ["proyectado"], ["propuesta"], ["work"]]) \
             or find_col(ws, sup, sub, ["ingresos","total"], [["proy"], ["proyectado"], ["propuesta"], ["work"]])

    # Componentes (para fallback)
    comp = {
        "acc_real": find_col(ws, sup, sub, ["ingreso","acceso"], [["real"]]),
        "trx_real": find_col(ws, sup, sub, ["ingreso","trans"],  [["real"]]),
        "per_real": find_col(ws, sup, sub, ["ingreso","perfil"], [["real"]]),
        "acc_proy": find_col(ws, sup, sub, ["ingreso","acceso"], [["proy"],["proyectado"],["propuesta"],["work"]]),
        "trx_proy": find_col(ws, sup, sub, ["ingreso","trans"],  [["proy"],["proyectado"],["propuesta"],["work"]]),
        "per_proy": find_col(ws, sup, sub, ["ingreso","perfil"], [["proy"],["proyectado"],["propuesta"],["work"]]),
    }
    return {"header": h, "cliente": cliente_c, "monto": monto_c, "real_total": real_c, "proy_total": proy_c, **comp}

# --------------------------------------
# Engines to read A.3 values
# --------------------------------------
def eval_with_model(model, sheet_name, addr):
    if not addr: return 0.0
    try:
        return safe_float(model.evaluate(f"'{sheet_name}'!{addr}"), 0.0)
    except Exception:
        return 0.0

def read_values_from_ws(ws, addr):
    if not addr: return 0.0
    return safe_float(ws[addr].value, 0.0)

def xlcalc_engine(wb):
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    mc = ModelCompiler(); model = Model(mc.read_and_parse_archive(bio))
    ws = wb[NEG_SHEET]
    cols = locate_columns(ws); h = cols["header"]; cliente_c = cols["cliente"]
    bolsas_map = {"BCS":"Chile","BVL":"Perú","BVC":"Colombia"}
    rows_cli = []; rows_bol = []; current_pais = None
    # iterar filas
    r = h+2
    while r <= ws.max_row:
        name = ws.cell(r, cliente_c).value
        if name is None or str(name).strip()=="":
            r += 1; continue
        name_s = str(name).strip()
        if name_s in bolsas_map:
            current_pais = bolsas_map[name_s]
            # fila agregada
            real_addr = ws.cell(r, cols["real_total"] or cols["acc_real"] or cols["trx_real"] or cols["per_real"]).coordinate if (cols["real_total"] or cols["acc_real"] or cols["trx_real"] or cols["per_real"]) else None
            proy_addr = ws.cell(r, cols["proy_total"] or cols["acc_proy"] or cols["trx_proy"] or cols["per_proy"]).coordinate if (cols["proy_total"] or cols["acc_proy"] or cols["trx_proy"] or cols["per_proy"]) else None
            rows_bol.append({"Pais": current_pais, "Bolsa": name_s,
                             "Real Excel": eval_with_model(model, NEG_SHEET, real_addr),
                             "Proyectado Excel": eval_with_model(model, NEG_SHEET, proy_addr)})
        else:
            # por cliente
            # direcciones
            monto_addr = ws.cell(r, cols["monto"]).coordinate if cols["monto"] else None
            # total directo
            real_addr = ws.cell(r, cols["real_total"]).coordinate if cols["real_total"] else None
            proy_addr = ws.cell(r, cols["proy_total"]).coordinate if cols["proy_total"] else None
            # componentes
            acc_real_addr = ws.cell(r, cols["acc_real"]).coordinate if cols["acc_real"] else None
            trx_real_addr = ws.cell(r, cols["trx_real"]).coordinate if cols["trx_real"] else None
            per_real_addr = ws.cell(r, cols["per_real"]).coordinate if cols["per_real"] else None
            acc_proy_addr = ws.cell(r, cols["acc_proy"]).coordinate if cols["acc_proy"] else None
            trx_proy_addr = ws.cell(r, cols["trx_proy"]).coordinate if cols["trx_proy"] else None
            per_proy_addr = ws.cell(r, cols["per_proy"]).coordinate if cols["per_proy"] else None

            monto = eval_with_model(model, NEG_SHEET, monto_addr)
            real_t = eval_with_model(model, NEG_SHEET, real_addr)
            proy_t = eval_with_model(model, NEG_SHEET, proy_addr)
            # fallback por componentes
            if real_t == 0 and any([acc_real_addr,trx_real_addr,per_real_addr]):
                real_t = sum([eval_with_model(model, NEG_SHEET, x) for x in [acc_real_addr,trx_real_addr,per_real_addr] if x])
            if proy_t == 0 and any([acc_proy_addr,trx_proy_addr,per_proy_addr]):
                proy_t = sum([eval_with_model(model, NEG_SHEET, x) for x in [acc_proy_addr,trx_proy_addr,per_proy_addr] if x])

            rows_cli.append({"Pais": current_pais or "", "Cliente": name_s,
                             "Monto USD": monto, "Real Excel": real_t, "Proyectado Excel": proy_t})
        r += 1

    return pd.DataFrame(rows_cli), pd.DataFrame(rows_bol), {"engine":"xlcalculator","columns":cols}

def cached_engine(wb):
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    wb_cache = openpyxl.load_workbook(bio, data_only=True)
    ws = wb_cache[NEG_SHEET]
    cols = locate_columns(ws); h = cols["header"]; cliente_c = cols["cliente"]
    bolsas_map = {"BCS":"Chile","BVL":"Perú","BVC":"Colombia"}
    rows_cli = []; rows_bol = []; current_pais = None
    r = h+2
    while r <= ws.max_row:
        name = ws.cell(r, cliente_c).value
        if name is None or str(name).strip()=="":
            r += 1; continue
        name_s = str(name).strip()
        def addr(colindex): return ws.cell(r, colindex).coordinate if colindex else None
        if name_s in bolsas_map:
            real_addr = addr(cols["real_total"] or cols["acc_real"] or cols["trx_real"] or cols["per_real"])
            proy_addr = addr(cols["proy_total"] or cols["acc_proy"] or cols["trx_proy"] or cols["per_proy"])
            current_pais = bolsas_map[name_s]
            rows_bol.append({"Pais": current_pais, "Bolsa": name_s,
                             "Real Excel": read_values_from_ws(ws, real_addr),
                             "Proyectado Excel": read_values_from_ws(ws, proy_addr)})
        else:
            monto_addr = addr(cols["monto"])
            real_addr  = addr(cols["real_total"])
            proy_addr  = addr(cols["proy_total"])
            acc_r, trx_r, per_r = addr(cols["acc_real"]), addr(cols["trx_real"]), addr(cols["per_real"])
            acc_p, trx_p, per_p = addr(cols["acc_proy"]), addr(cols["trx_proy"]), addr(cols["per_proy"])

            monto = read_values_from_ws(ws, monto_addr)
            real_t = read_values_from_ws(ws, real_addr)
            proy_t = read_values_from_ws(ws, proy_addr)
            if real_t == 0 and any([acc_r,trx_r,per_r]):
                real_t = sum(read_values_from_ws(ws, a) for a in [acc_r,trx_r,per_r] if a)
            if proy_t == 0 and any([acc_p,trx_p,per_p]):
                proy_t = sum(read_values_from_ws(ws, a) for a in [acc_p,trx_p,per_p] if a)

            rows_cli.append({"Pais": current_pais or "", "Cliente": name_s,
                             "Monto USD": monto, "Real Excel": real_t, "Proyectado Excel": proy_t})
        r += 1
    return pd.DataFrame(rows_cli), pd.DataFrame(rows_bol), {"engine":"cached","columns":cols}

def recalc_and_extract(wb):
    # 1) xlcalculator
    if XL_OK:
        try:
            return xlcalc_engine(wb)
        except Exception as e:
            meta1 = {"engine":"xlcalculator_failed", "error": str(e)}
    else:
        meta1 = {"engine":"xlcalculator_unavailable"}
    # 2) cached values
    try:
        df1, df2, meta2 = cached_engine(wb)
        meta2.update(meta1)
        return df1, df2, meta2
    except Exception as e:
        meta2 = {"engine":"fallback", "error": str(e)}
        meta2.update(meta1)
        return pd.DataFrame(columns=["Pais","Cliente","Monto USD","Real Excel","Proyectado Excel"]), \
               pd.DataFrame(columns=["Pais","Bolsa","Real Excel","Proyectado Excel"]), meta2

# --------------------------------------
# UI
# --------------------------------------
st.title("📊 Simulador Tarifario RV – Excel‑driven (v7)")
st.caption("Mueve los **tramos** en *1. Parametros* (columna **R+**) desde la barra lateral; "
           "la app escribe en el XLSX, **recalcula** (si hay motor) y lee **A.3 Negociación**. "
           "Si no aparece 'Ingreso Total', calculo el total como **Acceso + Transacción + Perfiles** por fila.")

uploaded = st.file_uploader("📁 Sube tu Excel maestro", type=["xlsx"])
if not uploaded:
    st.stop()

raw_bytes = uploaded.read()
wb_prefill = load_workbook(raw_bytes)
params0 = read_params_from_wb(wb_prefill)

with st.sidebar:
    st.header("⚙️ Parámetros (columna R+)")
    desc_bcs = st.number_input("Descuento BCS (0–1)", min_value=0.0, max_value=1.0, step=0.01, value=float(params0.get("desc_bcs",0.15)))

    st.subheader("Transacción – Tramos")
    trans_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["transaccion"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","bps","fijo"])
        df_show = df_e.rename(columns={"min":"Mín (USD)","max":"Máx (USD)","bps":"Variable (fracción/%)","fijo":"Fijo (USD)"})
        df_show = st.data_editor(df_show, key=f"trx_{pais}", use_container_width=True, num_rows="dynamic")
        rows = []
        for _, r in df_show.iterrows():
            rows.append({"min":safe_float(r.get("Mín (USD)",0)), "max":safe_float(r.get("Máx (USD)",float("inf"))),
                         "bps":safe_float(r.get("Variable (fracción/%)",0)), "fijo":safe_float(r.get("Fijo (USD)",0))})
        trans_edit[pais] = rows

    st.subheader("DMA – Tramos")
    dma_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["dma"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","bps","fijo"])
        df_show = df_e.rename(columns={"min":"Mín (USD)","max":"Máx (USD)","bps":"Variable (fracción/%)","fijo":"Fijo (USD)"})
        df_show = st.data_editor(df_show, key=f"dma_{pais}", use_container_width=True, num_rows="dynamic")
        rows = []
        for _, r in df_show.iterrows():
            rows.append({"min":safe_float(r.get("Mín (USD)",0)), "max":safe_float(r.get("Máx (USD)",float("inf"))),
                         "bps":safe_float(r.get("Variable (fracción/%)",0)), "fijo":safe_float(r.get("Fijo (USD)",0))})
        dma_edit[pais] = rows

    st.subheader("Acceso – Códigos/Pantallas")
    pant_edit = {}
    for pais in ["Chile","Colombia","Perú"]:
        base = params0["pantallas"].get(pais, [])
        df_e = pd.DataFrame(base) if base else pd.DataFrame(columns=["min","max","var","fija"])
        df_show = df_e.rename(columns={"min":"Mín #Códigos","max":"Máx #Códigos","var":"Variable por código (USD)","fija":"Fijo mensual tramo (USD)"})
        df_show = st.data_editor(df_show, key=f"pant_{pais}", use_container_width=True, num_rows="dynamic")
        rows = []
        for _, r in df_show.iterrows():
            rows.append({"min":safe_float(r.get("Mín #Códigos",0)), "max":safe_float(r.get("Máx #Códigos",float("inf"))),
                         "var":safe_float(r.get("Variable por código (USD)",0)), "fija":safe_float(r.get("Fijo mensual tramo (USD)",0))})
        pant_edit[pais] = rows

params_live = {"desc_bcs":desc_bcs, "transaccion":trans_edit, "dma":dma_edit, "pantallas":pant_edit}

with st.spinner("Aplicando parámetros y obteniendo resultados…"):
    wb = load_workbook(raw_bytes)
    wb = write_params_to_wb(wb, params_live)
    df_cli, df_bol, meta = recalc_and_extract(wb)

# KPIs y filtros
paises = ["Todos"] + sorted([p for p in df_cli["Pais"].dropna().unique().tolist() if isinstance(p,str) and p.strip()])
c1, c2 = st.columns(2)
with c1:
    pais_sel = st.selectbox("Filtrar por País", paises, index=0)
with c2:
    ver_detalle = st.toggle("Mostrar detalle por cliente", value=True)

df_f = df_cli if pais_sel=="Todos" else df_cli[df_cli["Pais"]==pais_sel]
tot_real = safe_float(df_f["Real Excel"].sum(),0.0)
tot_proy = safe_float(df_f["Proyectado Excel"].sum(),0.0)
monto    = safe_float(df_f["Monto USD"].sum(),0.0)

a,b,c,d = st.columns(4)
a.metric("Ingreso Real (Excel)", f"${tot_real:,.0f}")
b.metric("Ingreso Proyectado (Excel)", f"${tot_proy:,.0f}", delta=(f"{(tot_proy/tot_real-1)*100:+.1f}%" if tot_real>0 else None))
b.caption(f"Motor: **{meta.get('engine','?')}**")
c.metric("BPS Proyectado (Excel)", f"{calc_bps(tot_proy, monto):.2f} bps")
d.metric("Filas cargadas", f"{len(df_f):,}")

st.markdown("---")
cA, cB = st.columns(2, gap="large")
with cA:
    st.subheader("Real vs Proyectado – Excel")
    fig = go.Figure(data=[
        go.Bar(name="Real (Excel)", x=["Negociación"], y=[tot_real]),
        go.Bar(name="Proyectado (Excel)", x=["Negociación"], y=[tot_proy])
    ])
    fig.update_layout(barmode="group", height=300)
    st.plotly_chart(fig, use_container_width=True)
with cB:
    st.subheader("Totales por Bolsa (Excel)")
    st.dataframe(df_bol, use_container_width=True, height=300)

st.subheader("📋 Detalle por Cliente (Excel)")
if ver_detalle:
    tmp = df_f.copy()
    tmp["BPS"] = tmp.apply(lambda r: calc_bps(r["Proyectado Excel"], r["Monto USD"]), axis=1)
    st.dataframe(tmp.sort_values(["Pais","Proyectado Excel"], ascending=[True, False]), use_container_width=True, height=420)

st.markdown("---")
with st.expander("🔎 Diagnóstico"):
    st.json(meta)

col1, col2 = st.columns(2)
with col1:
    st.download_button("📥 Descargar Detalle (CSV)", df_f.to_csv(index=False).encode("utf-8"), "detalle_excel.csv", "text/csv", use_container_width=True)
with col2:
    out = BytesIO(); wb.save(out)
    st.download_button("📥 Descargar Excel (con parámetros escritos)", out.getvalue(), "excel_parametrizado.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
